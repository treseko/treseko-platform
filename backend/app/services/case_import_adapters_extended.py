from __future__ import annotations

import io
import json
from typing import Any

from .case_import_adapters import AdapterError, AdapterResult, _decode_text, _get, _key, _paired_steps, _stable_id

def _rows_from_matrix(matrix: list[list[Any]]) -> tuple[list[dict[str, str]], list[str]]:
    while matrix and not any(str(value or "").strip() for value in matrix[0]):
        matrix.pop(0)
    if not matrix:
        raise AdapterError("La planilla no contiene encabezados")
    raw_headers = matrix.pop(0)
    headers: list[str] = []
    counts: dict[str, int] = {}
    for position, raw in enumerate(raw_headers, 1):
        base = _key(raw) or f"column_{position}"
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    rows = []
    for raw_row in matrix:
        if not any(str(value or "").strip() for value in raw_row):
            continue
        padded = raw_row + [None] * max(0, len(headers) - len(raw_row))
        rows.append({headers[index]: str(padded[index] if padded[index] is not None else "").strip() for index in range(len(headers))})
    return rows, headers


def read_excel(data: bytes) -> tuple[list[dict[str, str]], list[str], list[str]]:
    matrix: list[list[Any]]
    warnings: list[str] = []
    if data.startswith(b"PK"):
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
            matrix = [list(row) for row in worksheet.iter_rows(values_only=True)]
            if len(workbook.sheetnames) > 1:
                warnings.append(f"La planilla contiene {len(workbook.sheetnames)} hojas; se importó solamente {worksheet.title}.")
            workbook.close()
        except Exception as exc:
            raise AdapterError("El archivo XLSX no se pudo leer") from exc
    elif data.startswith(b"\xd0\xcf\x11\xe0"):
        try:
            import xlrd
            workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
            worksheet = workbook.sheet_by_index(0)
            matrix = [worksheet.row_values(index) for index in range(worksheet.nrows)]
            if workbook.nsheets > 1:
                warnings.append(f"La planilla contiene {workbook.nsheets} hojas; se importó solamente {worksheet.name}.")
            workbook.release_resources()
        except Exception as exc:
            raise AdapterError("El archivo XLS legado no se pudo leer") from exc
    else:
        raise AdapterError("El archivo qTest debe ser una planilla XLSX o XLS válida")
    rows, headers = _rows_from_matrix(matrix)
    return rows, headers, warnings


def parse_qtest_excel(data: bytes) -> AdapterResult:
    rows, headers, warnings = read_excel(data)
    cases: list[dict[str, Any]] = []
    grouped: dict[str, dict[str, Any]] = {}
    current: dict[str, Any] | None = None
    attachment_fields = [field for field in headers if "attachment" in field]
    for position, row in enumerate(rows, 1):
        title = _get(row, "test case name", "testcase name", "name", "title")
        external_id = _get(row, "test case id", "testcase id", "id")
        module = _get(row, "module", "module path", "folder") or "Importados/qTest"
        identity = external_id or (f"{module}|{title}" if title else "")
        if identity and identity in grouped:
            current = grouped[identity]
        elif title:
            current = {
                "id": external_id or _stable_id("qtest", module, title),
                "external_version": _get(row, "version", "test case version"),
                "title": title,
                "description": _get(row, "description", "test case description"),
                "preconditions": _get(row, "precondition", "preconditions"),
                "priority": _get(row, "priority"),
                "status": _get(row, "status"),
                "type": _get(row, "type", "test type"),
                "suite_path": module.replace("\\", "/"),
                "steps": [],
            }
            cases.append(current)
            grouped[identity or current["id"]] = current
        elif current is None:
            warnings.append(f"Se omitió la fila qTest {position}: no se pudo asociar a un caso.")
            continue
        action = _get(row, "test step description", "step description", "test step", "step", "action")
        if action:
            current["steps"].append({
                "accion": action,
                "datos": _get(row, "test data", "step data", "data") or None,
                "resultado_esperado": _get(row, "expected result", "expected") or "Resultado esperado",
            })
    if attachment_fields and any(any(row.get(field) for field in attachment_fields) for row in rows):
        warnings.append("La planilla qTest contiene referencias a adjuntos; los binarios deben exportarse por separado.")
    return AdapterResult(cases, warnings=warnings, source_fields=headers, ignored_fields=attachment_fields)


def _nested_name(value: Any) -> str:
    if isinstance(value, dict):
        return str(value.get("name") or value.get("value") or "").strip()
    return str(value or "").strip()


def parse_zephyr_json(data: bytes) -> AdapterResult:
    try:
        parsed = json.loads(data.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise AdapterError("El JSON de Zephyr no es válido") from exc
    if isinstance(parsed, dict):
        values = parsed.get("values") or parsed.get("testCases") or parsed.get("items")
    else:
        values = parsed
    if not isinstance(values, list):
        raise AdapterError("El JSON de Zephyr no contiene una lista o un envoltorio values compatible")
    cases: list[dict[str, Any]] = []
    warnings: list[str] = []
    for position, item in enumerate(values, 1):
        if not isinstance(item, dict):
            warnings.append(f"Se omitió el elemento Zephyr {position}: no es un objeto.")
            continue
        folder = item.get("folder")
        folder_name = _nested_name(folder)
        if not folder_name and isinstance(folder, dict) and folder.get("id"):
            folder_name = f"Carpeta {folder['id']}"
        if not folder_name and item.get("folderId"):
            folder_name = f"Carpeta {item['folderId']}"
        script = item.get("testScript") if isinstance(item.get("testScript"), dict) else {}
        script_steps = script.get("steps") or item.get("steps") or []
        steps = []
        if isinstance(script_steps, list):
            for step in script_steps:
                if not isinstance(step, dict):
                    continue
                action = step.get("description") or step.get("action") or step.get("step")
                if action:
                    steps.append({"accion": str(action), "datos": step.get("testData") or step.get("data"), "resultado_esperado": step.get("expectedResult") or step.get("expected") or "Resultado esperado"})
        if script.get("self") and not steps:
            warnings.append("El JSON referencia scripts de Zephyr por URL; exportá también los scripts para conservar los pasos.")
        cases.append({
            "id": item.get("key") or item.get("id") or _stable_id("zephyr", item.get("name"), position),
            "title": item.get("name") or item.get("title"),
            "description": item.get("objective") or item.get("description"),
            "preconditions": item.get("precondition") or item.get("preconditions"),
            "priority": _nested_name(item.get("priority")) or item.get("priorityName"),
            "status": _nested_name(item.get("status")) or item.get("statusName"),
            "tags": item.get("labels") or [],
            "suite_path": f"Importados/Zephyr/{folder_name}" if folder_name else "Importados/Zephyr",
            "steps": steps,
        })
    return AdapterResult(cases, warnings=list(dict.fromkeys(warnings)), source_fields=sorted({str(key) for item in values if isinstance(item, dict) for key in item}))


def parse_gherkin(data: bytes) -> AdapterResult:
    try:
        from gherkin.parser import Parser
        from gherkin.pickles.compiler import Compiler
    except ImportError as exc:
        raise AdapterError("Falta la dependencia oficial gherkin-official") from exc
    text, warnings = _decode_text(data)
    try:
        document = Parser().parse(text)
    except Exception as exc:
        raise AdapterError(f"El archivo Gherkin no es válido: {exc}") from exc
    feature = document.get("feature") or {}
    if not feature:
        raise AdapterError("El archivo no contiene una Feature Gherkin")
    document["uri"] = "import.feature"
    pickles = Compiler().compile(document)
    cases: list[dict[str, Any]] = []
    for position, pickle in enumerate(pickles, 1):
        steps = []
        for step in pickle.get("steps", []):
            prefix = {"Context": "Given", "Action": "When", "Outcome": "Then", "Conjunction": "And"}.get(step.get("type"), "*")
            argument = step.get("argument")
            data_value = json.dumps(argument, ensure_ascii=False) if argument else None
            steps.append({"accion": f"{prefix} {step.get('text', '')}".strip(), "datos": data_value, "resultado_esperado": "Resultado esperado"})
        cases.append({
            "id": _stable_id("gherkin", pickle.get("uri"), pickle.get("name"), pickle.get("location", {}).get("line"), position),
            "title": pickle.get("name") or f"Escenario línea {pickle.get('location', {}).get('line', position)}",
            "description": feature.get("description") or None,
            "suite_path": feature.get("name") or "Importados/Gherkin",
            "tags": [tag.get("name") for tag in pickle.get("tags", []) if tag.get("name")],
            "steps": steps,
        })
    if not cases:
        warnings.append("La Feature es válida pero no genera escenarios ejecutables.")
    return AdapterResult(cases, warnings=warnings, source_fields=["feature", "scenario", "background", "rule", "examples", "steps"])
